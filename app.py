from flask import Flask, render_template, request, jsonify, send_file
import os
import uuid
import pandas as pd
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['PREPROCESSED_FOLDER'] = os.path.join(os.path.dirname(__file__), 'static', 'preprocessed')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB limit

# Ensure folders exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PREPROCESSED_FOLDER'], exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload_processed', methods=['POST'])
def upload_processed():
    """
    Saves the preprocessed image uploaded from client-side canvas.
    This is used so that the user can preview the optimized image in the modal.
    """
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "Không tìm thấy file"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "Tên file rỗng"}), 400
        
    unique_id = str(uuid.uuid4())
    filename = f"proc_{unique_id}.jpg"
    
    file_path = os.path.join(app.config['PREPROCESSED_FOLDER'], filename)
    file.save(file_path)
    
    return jsonify({
        "status": "success",
        "preprocessed_url": f"/static/preprocessed/{filename}"
    })

@app.route('/export', methods=['POST'])
def export_excel():
    try:
        # Get data from request body (JSON array)
        data = request.json
        if not data or not isinstance(data, list):
            return jsonify({"status": "error", "message": "Dữ liệu không hợp lệ"}), 400
            
        # Structure the data for pandas
        # Expected columns: STT, Chủ sở hữu, Địa chỉ, Diện tích, Mục đích sử dụng đất
        rows = []
        for i, item in enumerate(data):
            rows.append({
                "STT": item.get("stt", i + 1),
                "Chủ sở hữu": item.get("owner", ""),
                "Địa chỉ": item.get("address", ""),
                "Diện tích": item.get("area", ""),
                "Mục đích sử dụng đất": item.get("purpose", "")
            })
            
        df = pd.DataFrame(rows)
        
        # Write to temporary excel file
        temp_filename = f"export_{uuid.uuid4().hex}.xlsx"
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
        
        # Export using pandas and openpyxl
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Danh sách đất đai')
            
            # Format header row
            workbook = writer.book
            worksheet = writer.sheets['Danh sách đất đai']
            
            # Format header row
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            header_font = Font(name='Manrope', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='004C7E', end_color='004C7E', fill_type='solid') # Sacombank Dark Blue
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='C0C7D2'),
                right=Side(style='thin', color='C0C7D2'),
                top=Side(style='thin', color='C0C7D2'),
                bottom=Side(style='thin', color='C0C7D2')
            )
            
            # Format headers
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Adjust column width and add borders/alignment to rows
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    # Apply border to all cells
                    cell.border = thin_border
                    if cell.row > 1:
                        # Alignment for data
                        if cell.column == 1: # STT
                            cell.alignment = Alignment(horizontal='center')
                        elif cell.column == 4: # Area
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='left')
                            
                    val = str(cell.value or '')
                    if len(val) > max_len:
                        max_len = len(val)
                # Set width with safety padding
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            # Set header row height
            worksheet.row_dimensions[1].height = 25
            
        return send_file(
            temp_path,
            as_attachment=True,
            download_name='Danh_sach_trich_xuat_dat.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Lỗi khi xuất file Excel: {str(e)}"}), 500

if __name__ == '__main__':
    # Start the server locally on port 5000
    app.run(debug=True, host='0.0.0.0', port=5000)
